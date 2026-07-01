import json
from io import BytesIO

from openpyxl import load_workbook

from api.db.services.document_compare_report_service import DocumentCompareReportService


def sample_document():
    return {
        "document_id": "doc-a",
        "filename": "law.md",
        "source_path": "/safe/law.md",
        "mime_type": "text/markdown",
        "paragraphs": [{"text": "甲方应在三十日内付款。"}],
        "lines": [{"text": "甲方应在三十日内付款。"}],
        "tables": [],
        "audit": {"audit_id": "audit-1", "path": "/safe/law.md", "action": "read", "allowed": True},
    }


def sample_conflict():
    return {
        "conflict_id": "conflict-1",
        "severity": "high",
        "reason_code": "deadline_longer_than_allowed",
        "reason": "Target deadline 90 days is longer than standard deadline 30 days.",
        "standard_item": {"text": "甲方应在三十日内付款。"},
        "target_item": {"text": "甲方应在90日内付款。"},
        "evidence": {
            "left": {"source_ref": "law.md | line 1"},
            "right": {"source_ref": "contract.md | line 1"},
        },
    }


def test_document_compare_report_service_builds_auditable_payload_and_markdown():
    report = DocumentCompareReportService.build_report(
        title="合同法律冲突报告",
        documents=[sample_document()],
        diff={"summary": {"replace": 1, "insert": 0}},
        matches={"matches": [{"relation": "ambiguous", "score": 0.8, "left_item": {"text": "A"}, "right_item": {"text": "B"}}]},
        conflicts={"conflicts": [sample_conflict()]},
        run_id="run-1",
        agent_id="agent-1",
    )
    markdown = DocumentCompareReportService.render_markdown(report)

    assert report["risk_level"] == "high"
    assert report["audit"]["file_access"][0]["audit_id"] == "audit-1"
    assert "law.md | line 1" in {item["source_ref"] for item in report["references"]}
    assert "# 合同法律冲突报告" in markdown
    assert "run-1" in markdown
    assert "Target deadline 90 days" in markdown


def test_document_compare_report_service_renders_json_docx_and_xlsx():
    report = DocumentCompareReportService.build_report(
        documents=[sample_document()],
        diff={"summary": {"replace": 1}},
        conflicts={"conflicts": [sample_conflict()]},
        run_id="run-1",
    )

    json_bytes, json_mime = DocumentCompareReportService.render_bytes(report, "json")
    docx_bytes, docx_mime = DocumentCompareReportService.render_bytes(report, "docx")
    xlsx_bytes, xlsx_mime = DocumentCompareReportService.render_bytes(report, "xlsx")

    assert json.loads(json_bytes.decode("utf-8"))["run_id"] == "run-1"
    assert json_mime == "application/json"
    assert docx_mime.endswith("wordprocessingml.document")
    assert docx_bytes.startswith(b"PK")
    assert xlsx_mime.endswith("spreadsheetml.sheet")
    workbook = load_workbook(BytesIO(xlsx_bytes))
    assert {"Summary", "Files", "Diff", "Matches", "Conflicts", "Missing", "Audit"}.issubset(set(workbook.sheetnames))
