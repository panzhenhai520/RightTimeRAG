#
#  Copyright 2026 The InfiniFlow Authors. All Rights Reserved.
#
#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

from __future__ import annotations

import csv
import fnmatch
import hashlib
import mimetypes
import os
import re
import time
from pathlib import Path
from typing import Any

from common.misc_utils import get_uuid


class WorkspaceFileError(Exception):
    """Structured error for safe workspace file access."""

    def __init__(self, code: str, message: str, details: dict[str, Any] | None = None):
        super().__init__(message)
        self.code = code
        self.details = details or {}

    def to_dict(self) -> dict[str, Any]:
        return {"error_code": self.code, "message": str(self), "details": self.details}


class WorkspaceFileService:
    """Read-only workspace file access behind an explicit root allowlist."""

    ROOTS_ENV = "AGENT_WORKSPACE_ROOTS"
    ROOT_ENV = "AGENT_WORKSPACE_ROOT"
    AGENT_WORKSPACE_BASE_ENV = "AGENT_WORKSPACE_BASE"
    DEFAULT_AGENT_WORKSPACE_DIR = "agent_workspaces"
    DEFAULT_MAX_RESULTS = 500
    DEFAULT_MAX_READ_BYTES = 256 * 1024
    DEFAULT_MAX_TABLE_ROWS = 500
    DEFAULT_MAX_TABLE_CELLS = 10000

    @classmethod
    def configured_roots(cls) -> list[Path]:
        raw_roots = []
        if os.environ.get(cls.ROOTS_ENV):
            raw_roots.extend(os.environ.get(cls.ROOTS_ENV, "").split(os.pathsep))
        if os.environ.get(cls.ROOT_ENV):
            raw_roots.append(os.environ.get(cls.ROOT_ENV, ""))
        if not raw_roots:
            raw_roots.append(os.getcwd())

        roots = []
        seen = set()
        for raw in raw_roots:
            if not str(raw or "").strip():
                continue
            root = Path(str(raw)).expanduser().resolve()
            if not root.exists() or not root.is_dir():
                continue
            key = str(root)
            if key in seen:
                continue
            seen.add(key)
            roots.append(root)
        if not roots:
            raise WorkspaceFileError("NO_WORKSPACE_ROOT", "No accessible workspace root is configured.")
        return roots

    @classmethod
    def list_roots(cls, *, roots: list[str | Path] | None = None, agent_id: str = "") -> list[dict[str, Any]]:
        return [cls._root_info(index, root) for index, root in enumerate(cls._roots(roots, agent_id=agent_id))]

    @classmethod
    def agent_workspace_base(cls) -> Path:
        raw = str(os.environ.get(cls.AGENT_WORKSPACE_BASE_ENV) or "").strip()
        base = Path(raw).expanduser() if raw else Path(os.getcwd()) / cls.DEFAULT_AGENT_WORKSPACE_DIR
        base = base.resolve(strict=False)
        base.mkdir(parents=True, exist_ok=True)
        return base

    @staticmethod
    def _safe_workspace_name(agent_id: str) -> str:
        safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(agent_id or "").strip())
        return safe.strip("._-") or "default"

    @classmethod
    def ensure_agent_workspace(cls, agent_id: str, *, title: str = "", tenant_id: str = "") -> dict[str, Any]:
        workspace = cls.agent_workspace_base() / cls._safe_workspace_name(agent_id)
        workspace.mkdir(parents=True, exist_ok=True)
        for child in ("input", "output", "tmp"):
            (workspace / child).mkdir(parents=True, exist_ok=True)
        readme = workspace / "README.md"
        if not readme.exists():
            readme.write_text(
                "\n".join(
                    [
                        "# Agent Workspace",
                        "",
                        "This directory is the managed file space for one RAGFlow agent.",
                        "The agent has full read/write permission inside this directory only.",
                        "",
                    ]
                ),
                encoding="utf-8",
            )
        default_report = workspace / "output" / "report.md"
        if not default_report.exists():
            default_report.write_text("# Agent Report Output\n", encoding="utf-8")
        root_info = cls._root_info(0, workspace)
        return {
            "schema_version": 1,
            "managed": True,
            "agent_id": str(agent_id or ""),
            "tenant_id": str(tenant_id or ""),
            "title": str(title or ""),
            "root": root_info,
            "root_path": str(workspace),
            "input_path": "input",
            "output_path": "output",
            "tmp_path": "tmp",
            "default_output_file": "output/report.md",
            "permissions": {
                "read": True,
                "write": True,
                "create": True,
                "overwrite": True,
                "append": True,
                "delete": False,
                "patch": True,
                "scope": "agent_workspace",
            },
        }

    @classmethod
    def agent_workspace_roots(cls, agent_id: str) -> list[Path]:
        if not str(agent_id or "").strip():
            return []
        return [Path(cls.ensure_agent_workspace(agent_id)["root_path"])]

    @classmethod
    def stat(
        cls,
        *,
        path: str = ".",
        root: str = "",
        roots: list[str | Path] | None = None,
        tenant_id: str = "",
        user_id: str = "",
        run_id: str = "",
    ) -> dict[str, Any]:
        resolved, root_info = cls.resolve(path=path, root=root, roots=roots, must_exist=True)
        audit = cls.audit_record(
            action="stat",
            tenant_id=tenant_id,
            user_id=user_id,
            run_id=run_id,
            path=str(resolved),
            allowed=True,
        )
        return {"file": cls.file_info(resolved, root_info), "audit": audit}

    @classmethod
    def list_files(
        cls,
        *,
        path: str = ".",
        root: str = "",
        roots: list[str | Path] | None = None,
        recursive: bool = False,
        include_dirs: bool = True,
        extensions: list[str] | str | None = None,
        pattern: str = "",
        regex: str = "",
        max_results: int | None = None,
        tenant_id: str = "",
        user_id: str = "",
        run_id: str = "",
    ) -> dict[str, Any]:
        base, root_info = cls.resolve(path=path, root=root, roots=roots, must_exist=True)
        if not base.is_dir():
            raise WorkspaceFileError("NOT_A_DIRECTORY", "Workspace path is not a directory.", {"path": str(base)})
        entries = []
        limit = cls._limit(max_results, cls.DEFAULT_MAX_RESULTS)
        iterator = base.rglob("*") if recursive else base.iterdir()
        for item in iterator:
            if not include_dirs and item.is_dir():
                continue
            if not cls._matches(item, extensions=extensions, pattern=pattern, regex=regex):
                continue
            entries.append(cls.file_info(item, root_info))
            if len(entries) >= limit:
                break
        return {
            "root": root_info,
            "path": str(base),
            "files": entries,
            "count": len(entries),
            "truncated": len(entries) >= limit,
            "audit": cls.audit_record("list", tenant_id=tenant_id, user_id=user_id, run_id=run_id, path=str(base), allowed=True),
        }

    @classmethod
    def search_files(
        cls,
        *,
        query: str = "",
        path: str = ".",
        root: str = "",
        roots: list[str | Path] | None = None,
        extensions: list[str] | str | None = None,
        pattern: str = "",
        regex: str = "",
        max_results: int | None = None,
        tenant_id: str = "",
        user_id: str = "",
        run_id: str = "",
    ) -> dict[str, Any]:
        base, root_info = cls.resolve(path=path, root=root, roots=roots, must_exist=True)
        if not base.is_dir():
            raise WorkspaceFileError("NOT_A_DIRECTORY", "Workspace search path is not a directory.", {"path": str(base)})
        normalized_query = str(query or "").strip().lower()
        entries = []
        limit = cls._limit(max_results, cls.DEFAULT_MAX_RESULTS)
        for item in base.rglob("*"):
            if item.is_dir():
                continue
            if normalized_query and normalized_query not in item.name.lower() and normalized_query not in str(item.relative_to(root_info["path"])).lower():
                continue
            if not cls._matches(item, extensions=extensions, pattern=pattern, regex=regex):
                continue
            entries.append(cls.file_info(item, root_info))
            if len(entries) >= limit:
                break
        return {
            "root": root_info,
            "path": str(base),
            "query": query or "",
            "files": entries,
            "count": len(entries),
            "truncated": len(entries) >= limit,
            "audit": cls.audit_record("search", tenant_id=tenant_id, user_id=user_id, run_id=run_id, path=str(base), allowed=True),
        }

    @classmethod
    def read_file(
        cls,
        *,
        path: str,
        root: str = "",
        roots: list[str | Path] | None = None,
        encoding: str = "utf-8",
        max_bytes: int | None = None,
        tenant_id: str = "",
        user_id: str = "",
        run_id: str = "",
    ) -> dict[str, Any]:
        resolved, root_info = cls.resolve(path=path, root=root, roots=roots, must_exist=True)
        cls._require_file(resolved)
        limit = cls._limit(max_bytes, cls.DEFAULT_MAX_READ_BYTES)
        with resolved.open("rb") as file:
            payload = file.read(limit + 1)
        truncated = len(payload) > limit
        if truncated:
            payload = payload[:limit]
        text = payload.decode(encoding or "utf-8", errors="replace")
        return {
            "file": cls.file_info(resolved, root_info),
            "content": text,
            "encoding": encoding or "utf-8",
            "bytes_read": len(payload),
            "truncated": truncated,
            "line_count": text.count("\n") + (1 if text else 0),
            "source_ref": cls.source_ref(resolved, root_info),
            "audit": cls.audit_record("read", tenant_id=tenant_id, user_id=user_id, run_id=run_id, path=str(resolved), allowed=True),
        }

    @classmethod
    def read_range(
        cls,
        *,
        path: str,
        root: str = "",
        roots: list[str | Path] | None = None,
        start_line: int = 1,
        end_line: int | None = None,
        encoding: str = "utf-8",
        max_bytes: int | None = None,
        tenant_id: str = "",
        user_id: str = "",
        run_id: str = "",
    ) -> dict[str, Any]:
        resolved, root_info = cls.resolve(path=path, root=root, roots=roots, must_exist=True)
        cls._require_file(resolved)
        start_line = max(1, int(start_line or 1))
        end_line = int(end_line) if end_line is not None else start_line
        if end_line < start_line:
            raise WorkspaceFileError("INVALID_RANGE", "end_line must be greater than or equal to start_line.")
        limit = cls._limit(max_bytes, cls.DEFAULT_MAX_READ_BYTES)
        selected = []
        bytes_read = 0
        truncated = False
        with resolved.open("r", encoding=encoding or "utf-8", errors="replace") as file:
            for line_number, line in enumerate(file, start=1):
                if line_number < start_line:
                    continue
                if line_number > end_line:
                    break
                encoded_len = len(line.encode(encoding or "utf-8", errors="replace"))
                if bytes_read + encoded_len > limit:
                    truncated = True
                    break
                bytes_read += encoded_len
                selected.append({"line_number": line_number, "text": line.rstrip("\n")})
        content = "\n".join(item["text"] for item in selected)
        return {
            "file": cls.file_info(resolved, root_info),
            "lines": selected,
            "content": content,
            "line_start": start_line,
            "line_end": selected[-1]["line_number"] if selected else start_line,
            "bytes_read": bytes_read,
            "truncated": truncated,
            "source_ref": f"{cls.source_ref(resolved, root_info)} | lines {start_line}-{selected[-1]['line_number'] if selected else start_line}",
            "audit": cls.audit_record("read_range", tenant_id=tenant_id, user_id=user_id, run_id=run_id, path=str(resolved), allowed=True),
        }

    @classmethod
    def read_table(
        cls,
        *,
        path: str,
        root: str = "",
        roots: list[str | Path] | None = None,
        sheet_name: str = "",
        header_row: int = 1,
        start_row: int | None = None,
        max_rows: int | None = None,
        max_cells: int | None = None,
        encoding: str = "utf-8",
        tenant_id: str = "",
        user_id: str = "",
        run_id: str = "",
    ) -> dict[str, Any]:
        resolved, root_info = cls.resolve(path=path, root=root, roots=roots, must_exist=True)
        cls._require_file(resolved)
        suffix = resolved.suffix.lower()
        if suffix in {".csv", ".tsv", ".txt"}:
            table = cls._read_delimited_table(
                resolved,
                delimiter="\t" if suffix == ".tsv" else ",",
                encoding=encoding or "utf-8",
                max_rows=cls._limit(max_rows, cls.DEFAULT_MAX_TABLE_ROWS),
            )
        elif suffix == ".xlsx":
            table = cls._read_xlsx_table(
                resolved,
                sheet_name=sheet_name,
                header_row=max(1, int(header_row or 1)),
                start_row=start_row,
                max_rows=cls._limit(max_rows, cls.DEFAULT_MAX_TABLE_ROWS),
            )
        else:
            raise WorkspaceFileError("UNSUPPORTED_TABLE_FORMAT", f"Unsupported table file format: {suffix or 'unknown'}")

        cell_limit = cls._limit(max_cells, cls.DEFAULT_MAX_TABLE_CELLS)
        if len(table["rows"]) * max(1, len(table["headers"])) > cell_limit:
            keep_rows = max(0, cell_limit // max(1, len(table["headers"])))
            table["rows"] = table["rows"][:keep_rows]
            table["truncated"] = True
        table["file"] = cls.file_info(resolved, root_info)
        table["source_ref"] = f"{cls.source_ref(resolved, root_info)} | table"
        table["audit"] = cls.audit_record("read_table", tenant_id=tenant_id, user_id=user_id, run_id=run_id, path=str(resolved), allowed=True)
        return table

    @classmethod
    def resolve(
        cls,
        *,
        path: str,
        root: str = "",
        roots: list[str | Path] | None = None,
        must_exist: bool = True,
    ) -> tuple[Path, dict[str, Any]]:
        allowed_roots = cls._roots(roots)
        selected_roots = cls._select_roots(root, allowed_roots)
        raw_path = str(path or ".").strip() or "."
        candidate = Path(raw_path).expanduser()
        candidates = [candidate.resolve(strict=False)] if candidate.is_absolute() else [(base / candidate).resolve(strict=False) for base in selected_roots]
        for item in candidates:
            for index, allowed_root in enumerate(allowed_roots):
                if cls._is_relative_to(item, allowed_root):
                    if must_exist and not item.exists():
                        raise WorkspaceFileError("PATH_NOT_FOUND", "Workspace path does not exist.", {"path": str(item)})
                    return item, cls._root_info(index, allowed_root)
        raise WorkspaceFileError("PATH_OUTSIDE_ROOT", "Workspace path is outside configured roots.", {"path": raw_path})

    @classmethod
    def file_info(cls, path: Path, root_info: dict[str, Any]) -> dict[str, Any]:
        stat = path.stat()
        rel = str(path.relative_to(root_info["path"])) if cls._is_relative_to(path, Path(root_info["path"])) else path.name
        mime_type, _ = mimetypes.guess_type(path.name)
        return {
            "name": path.name,
            "path": str(path),
            "relative_path": rel,
            "root_id": root_info["root_id"],
            "type": "directory" if path.is_dir() else "file",
            "size": stat.st_size,
            "modified_at": stat.st_mtime,
            "mime_type": mime_type or "",
            "extension": path.suffix.lower(),
            "sha256": cls._sha256(path) if path.is_file() and stat.st_size <= cls.DEFAULT_MAX_READ_BYTES else "",
        }

    @staticmethod
    def source_ref(path: Path, root_info: dict[str, Any]) -> str:
        rel = str(path.relative_to(root_info["path"])) if WorkspaceFileService._is_relative_to(path, Path(root_info["path"])) else path.name
        return rel

    @staticmethod
    def audit_record(
        action: str,
        *,
        tenant_id: str = "",
        user_id: str = "",
        run_id: str = "",
        path: str = "",
        allowed: bool = True,
        reason: str = "",
    ) -> dict[str, Any]:
        return {
            "audit_id": get_uuid(),
            "created_at": time.time(),
            "action": action,
            "tenant_id": tenant_id or "",
            "user_id": user_id or "",
            "run_id": run_id or "",
            "path": path or "",
            "allowed": bool(allowed),
            "reason": reason or "",
        }

    @classmethod
    def _roots(cls, roots: list[str | Path] | None = None, *, agent_id: str = "") -> list[Path]:
        if roots is None and str(agent_id or "").strip():
            return cls.agent_workspace_roots(agent_id)
        if roots is None:
            return cls.configured_roots()
        normalized = []
        for root in roots:
            item = Path(root).expanduser().resolve()
            if item.exists() and item.is_dir():
                normalized.append(item)
        if not normalized:
            raise WorkspaceFileError("NO_WORKSPACE_ROOT", "No accessible workspace root is configured.")
        return normalized

    @staticmethod
    def _root_info(index: int, root: Path) -> dict[str, Any]:
        digest = hashlib.sha1(str(root).encode("utf-8")).hexdigest()[:12]
        return {"root_id": f"root-{index}-{digest}", "path": str(root), "name": root.name or str(root)}

    @staticmethod
    def _select_roots(root: str, roots: list[Path]) -> list[Path]:
        if not root:
            return roots
        root_text = str(root).strip()
        selected = []
        for index, item in enumerate(roots):
            info = WorkspaceFileService._root_info(index, item)
            if root_text in {info["root_id"], str(item), item.name}:
                selected.append(item)
        if not selected:
            raise WorkspaceFileError("ROOT_NOT_ALLOWED", "Requested workspace root is not configured.", {"root": root_text})
        return selected

    @staticmethod
    def _is_relative_to(path: Path, root: Path) -> bool:
        try:
            path.relative_to(root)
            return True
        except ValueError:
            return False

    @staticmethod
    def _limit(value: int | None, default: int) -> int:
        try:
            return max(1, min(int(value or default), default))
        except Exception:
            return default

    @staticmethod
    def _normalize_extensions(value: list[str] | str | None) -> set[str]:
        if not value:
            return set()
        items = value.split(",") if isinstance(value, str) else value
        result = set()
        for item in items:
            text = str(item or "").strip().lower()
            if not text:
                continue
            result.add(text if text.startswith(".") else f".{text}")
        return result

    @classmethod
    def _matches(cls, path: Path, *, extensions: list[str] | str | None = None, pattern: str = "", regex: str = "") -> bool:
        allowed_extensions = cls._normalize_extensions(extensions)
        if allowed_extensions and path.suffix.lower() not in allowed_extensions:
            return False
        if pattern and not fnmatch.fnmatch(path.name, pattern):
            return False
        if regex and not re.search(regex, str(path)):
            return False
        return True

    @staticmethod
    def _require_file(path: Path) -> None:
        if not path.is_file():
            raise WorkspaceFileError("NOT_A_FILE", "Workspace path is not a file.", {"path": str(path)})

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as file:
            for chunk in iter(lambda: file.read(65536), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _read_delimited_table(path: Path, *, delimiter: str, encoding: str, max_rows: int) -> dict[str, Any]:
        rows = []
        with path.open("r", encoding=encoding, errors="replace", newline="") as file:
            reader = csv.DictReader(file, delimiter=delimiter)
            headers = [str(item or "") for item in (reader.fieldnames or [])]
            for index, row in enumerate(reader, start=2):
                rows.append({"row_index": index, "values": dict(row)})
                if len(rows) >= max_rows:
                    break
        return {"headers": headers, "rows": rows, "row_count": len(rows), "truncated": len(rows) >= max_rows}

    @staticmethod
    def _read_xlsx_table(path: Path, *, sheet_name: str, header_row: int, start_row: int | None, max_rows: int) -> dict[str, Any]:
        try:
            import openpyxl
        except Exception as exc:
            raise WorkspaceFileError("XLSX_READER_UNAVAILABLE", "openpyxl is required to read xlsx files.") from exc
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        header_values = [cell.value for cell in next(sheet.iter_rows(min_row=header_row, max_row=header_row))]
        headers = [str(value or f"column_{index}") for index, value in enumerate(header_values, start=1)]
        rows = []
        first_data_row = int(start_row or header_row + 1)
        for row_index, cells in enumerate(sheet.iter_rows(min_row=first_data_row, values_only=True), start=first_data_row):
            values = {headers[index]: value for index, value in enumerate(cells[: len(headers)])}
            rows.append({"row_index": row_index, "values": values})
            if len(rows) >= max_rows:
                break
        return {"sheet_name": sheet.title, "headers": headers, "rows": rows, "row_count": len(rows), "truncated": len(rows) >= max_rows}
